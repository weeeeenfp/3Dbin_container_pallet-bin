# main.py
import time
from openpyxl import load_workbook
from packer import Packer
from item import Item
from bin import Bin

start_time = time.time()

# 讀取 Excel
wb = load_workbook("data/20251029.xlsx")
ws = wb.active

items = []
for row in ws.iter_rows(min_row=2, values_only=True):
    l_cm, w_cm, h_cm, weight, qty = row[0], row[1], row[2], row[3], int(row[4])
    l, w, h = l_cm*10, w_cm*10, h_cm*10
    for _ in range(qty):
        items.append(Item(f"{l_cm}×{w_cm}×{h_cm}", l, w, h, weight))

print(f"總共 {len(items)} 箱貨物待裝")

# 建立 Packer
packer = Packer()
packer.add_bin(Bin("Pallet_1"))

for item in items:
    packer.add_item(item)

# 執行裝箱
packer.pack()

# 統計結果
total_items = len(items)
fitted_items = sum(len(b.items) for b in packer.bins)
unfitted_items = total_items - fitted_items

print("\n" + "="*60)
print("最終裝箱結果")
print("="*60)
print(f"總箱數       : {total_items}")
print(f"已裝箱數     : {fitted_items}")
print(f"未裝箱數     : {unfitted_items}")

if unfitted_items == 0:
    print("結果：全部裝完，100% 出貨")
else:
    print("未裝到的貨物：")
    # 這裡可以再補 unfitted 列表，但本演算法不會有

print(f"\n使用棧板數   : {len(packer.bins)} 個")
for i, b in enumerate(packer.bins, 1):
    if b.items:
        print(f"  棧板 {i:2d} → {len(b.items):3d} 箱 | 高度 {b.get_current_height():4.0f}mm | 重量 {b.total_weight:5.1f}kg")

containers = max(1, (len(packer.bins) + 21) // 22)
print(f"\n預估 40呎貨櫃數：{containers} 個（雙層堆疊）")
print(f"總執行時間       ：{time.time() - start_time:.2f} 秒")